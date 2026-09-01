"""Power BI legacy/target source migration utility.

Inputs
------
1. Required mapping workbook with sheet "Column Mapping" and columns:
   Legacy Table, Target Table, Legacy Column Name, Legacy Data Type,
   Target Column Name, Target Data Type
2. Required PBIP semantic model definition/tables folder (read only)
3. Optional disabled Power Query export folder. Each original .txt file must use
   the exact query name. Generated _target/_legacy files share this folder.

Outputs
-------
* One full-table createOrReplace TMDL View script per changed table.
* One migrated TXT file per changed disabled Power Query export.
* One migration_log.xlsx workbook.

Hard safety invariant
---------------------
Model-facing table and column declaration names are immutable. Only native SQL,
M expressions, physical sourceColumn values, and physical-column dataType values
may change. Before writing a TMDL script, declaration names are restored from the
original PBIP definition by lineageTag and verified.
"""
from __future__ import annotations

import argparse
import hashlib
import os
import re
import sys
import tempfile
from collections import defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd

try:
    from sqlglot import exp, parse_one
    from sqlglot.errors import ParseError
    from sqlglot.optimizer.scope import build_scope
except ImportError:
    exp = None
    parse_one = None
    ParseError = Exception
    build_scope = None
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =============================================================================
# USER CONFIGURATION
# =============================================================================
DISABLED_POWER_QUERY_FOLDER: Optional[str] = r"" # Input Optional (Recommended to store this outside the pbip folder)
MAPPING_WORKBOOK: str = r"" # Input mapping workbook path (Recommended to store this outside the pbip folder)
MAPPING_SHEET: str = "Column Mapping" # Input sheet name 
TMDL_TABLES_INPUT_FOLDER: str = r"" # Input TMDL file path from the pbip folder
TMDL_OUTPUT_FOLDER: str = r"" # Input TMDL Output folder path from pbip folder
LOG_OUTPUT_FOLDER: str = r"" # Input a path to store log files (Recommended to store this outside the pbip folder)
DIRECTION: str = "target"              # target or legacy
DRY_RUN: bool = True
OVERWRITE_EXISTING_OUTPUTS: bool = True
DELETE_STALE_OUTPUTS: bool = False
TMDL_OUTPUT_MODE: str = "tmdl_view_script"  # tmdl_view_script or pbip_definition
LOG_FILENAME: str = "migration_log.xlsx"
LIVE_CONSOLE_PROGRESS: bool = True
M_CULTURE: str = "en-US"

# Default is destination type. Add only exceptions that must retain legacy type.
LEGACY_TYPE_RETENTION_RULES: list[dict[str, str]] = [
    {"legacy_type": "integer", "target_type": "text", "match_mode": "normalized",
     "description": "Retain legacy integer type when target is text"},
    {"legacy_type": "number", "target_type": "text", "match_mode": "normalized",
     "description": "Retain legacy floating numeric type when target is text"},
    {"legacy_type": "decimal", "target_type": "text", "match_mode": "normalized",
     "description": "Retain legacy decimal type when target is text"},
    {"legacy_type": "date", "target_type": "datetime", "match_mode": "normalized",
     "description": "Retain legacy date type when target includes time"},
]

COLUMN_TYPE_OVERRIDES: list[dict[str, str]] = [
    # {"legacy_table": "dwh.Table", "legacy_column": "Column",
    #  "final_type": "integer", "description": "Reason"}
]

# =============================================================================
# CONSTANTS AND TYPES
# =============================================================================
REQ = ["Legacy Table", "Target Table", "Legacy Column Name", "Legacy Data Type",
       "Target Column Name", "Target Data Type"]
GENERATED_SUFFIXES = ("_target", "_legacy")
TYPE_ALIASES = {
    "text":"text", "string":"text", "varchar":"text", "nvarchar":"text", "char":"text", "nchar":"text",
    "int":"integer", "integer":"integer", "bigint":"integer", "smallint":"integer", "tinyint":"integer",
    "int64":"integer", "whole number":"integer",
    "number":"number", "float":"number", "real":"number", "double":"number", "double precision":"number",
    "decimal":"decimal", "numeric":"decimal", "money":"decimal", "currency":"decimal", "smallmoney":"decimal",
    "date":"date", "datetime":"datetime", "datetime2":"datetime", "smalldatetime":"datetime", "timestamp":"datetime",
    "datetimezone":"datetimezone", "time":"time", "duration":"duration",
    "boolean":"boolean", "bool":"boolean", "logical":"boolean", "bit":"boolean",
    "binary":"binary", "varbinary":"binary", "any":"any",
}
M_TYPES = {"text":"type text", "integer":"Int64.Type", "number":"type number", "decimal":"type number",
           "date":"type date", "datetime":"type datetime", "datetimezone":"type datetimezone",
           "boolean":"type logical", "time":"type time", "duration":"type duration", "binary":"type binary", "any":"type any"}
TMDL_TYPES = {"text":"string", "integer":"int64", "number":"double", "decimal":"decimal",
              "date":"dateTime", "datetime":"dateTime", "datetimezone":"dateTime", "boolean":"boolean",
              "time":"dateTime", "duration":"double", "binary":"binary", "any":"string"}
SQL_TYPES = {"text":"VARCHAR(4000)", "integer":"BIGINT", "number":"FLOAT", "decimal":"DECIMAL(38,10)",
             "date":"DATE", "datetime":"DATETIME2", "datetimezone":"DATETIMEOFFSET", "boolean":"BIT",
             "time":"TIME", "binary":"VARBINARY(MAX)"}

@dataclass(frozen=True)
class TypeDecision:
    legacy_raw: str; target_raw: str; legacy_norm: str; target_norm: str
    final_norm: str; source: str; description: str = ""; retain_legacy: bool = False

@dataclass(frozen=True)
class ColMap:
    legacy_table: str; target_table: str; legacy_col: str; target_col: str
    legacy_type: str; target_type: str; decision: TypeDecision
    @property
    def src_table(self): return self.legacy_table if CFG.direction == "target" else self.target_table
    @property
    def dst_table(self): return self.target_table if CFG.direction == "target" else self.legacy_table
    @property
    def src_col(self): return self.legacy_col if CFG.direction == "target" else self.target_col
    @property
    def dst_col(self): return self.target_col if CFG.direction == "target" else self.legacy_col

@dataclass
class ModelMap:
    table_pairs: dict[str, tuple[str,str]]
    cols_by_src_table: dict[str,list[ColMap]]
    all_cols: list[ColMap]

@dataclass
class Config:
    pq_folder: Optional[Path]; workbook: Path; sheet: str; tmdl_in: Path; tmdl_out: Path
    log_out: Path; direction: str; dry_run: bool

@dataclass
class Changes:
    changed: bool=False; relevant: bool=False; classification: str="Unknown"
    table: list[dict]=field(default_factory=list); column: list[dict]=field(default_factory=list)
    dtype: list[dict]=field(default_factory=list); sql: list[dict]=field(default_factory=list)
    m: list[dict]=field(default_factory=list); protected: list[dict]=field(default_factory=list)
    warnings: list[dict]=field(default_factory=list); errors: list[dict]=field(default_factory=list)

CFG: Config
QUERY_CONTEXT: dict[str,set[str]] = {}

# =============================================================================
# UTILITIES
# =============================================================================
def clean(v: Any) -> str:
    if v is None or (isinstance(v,float) and pd.isna(v)): return ""
    return str(v).strip()

def key(v: str) -> str: return clean(v).casefold()
def raw_type(v: str) -> str: return re.sub(r"\s*\([^)]*\)\s*$", "", clean(v).casefold()).strip()
def norm_type(v: str) -> str: return TYPE_ALIASES.get(raw_type(v), raw_type(v))
def strip_quotes(v: str) -> str:
    v=clean(v)
    return v[1:-1].replace("''", "'") if len(v)>=2 and v[0]==v[-1]=="'" else v

def quote_like(old: str, new: str) -> str:
    return "'"+new.replace("'","''")+"'" if old.strip().startswith("'") or re.search(r"\s|[^A-Za-z0-9_]",new) else new

def split_table(v: str) -> tuple[str,str]:
    b=re.fullmatch(r"\[([^]]+)\]\.\[([^]]+)\]",clean(v))
    if b: return b.group(1),b.group(2)
    p=clean(v).strip("[]").split(".",1)
    if len(p)!=2: raise ValueError(f"Expected schema.object, found {v!r}")
    return p[0].strip('[]"'),p[1].strip('[]"')

def sha(s:str)->str: return hashlib.sha256(s.encode("utf-8")).hexdigest()
def read_text(p:Path)->tuple[str,str]:
    b=p.read_bytes(); enc="utf-8-sig" if b.startswith(b"\xef\xbb\xbf") else "utf-8"
    return b.decode(enc), "\r\n" if b"\r\n" in b else "\n"
def newline(text:str,nl:str)->str: return text.replace("\r\n","\n").replace("\r","\n").replace("\n",nl)
def atomic_write(p:Path,text:str):
    p.parent.mkdir(parents=True,exist_ok=True); tmp=None
    try:
        with tempfile.NamedTemporaryFile("w",encoding="utf-8",newline="",dir=p.parent,delete=False,suffix=".tmp") as f:
            f.write(text); f.flush(); os.fsync(f.fileno()); tmp=Path(f.name)
        os.replace(tmp,p); tmp=None
    finally:
        if tmp and tmp.exists(): tmp.unlink()

def relative(child:Path,parent:Path)->bool:
    try: child.relative_to(parent); return True
    except ValueError: return False

def sql_ident(v:str)->str: return "["+v.replace("]","]]" )+"]"

def build_config()->Config:
    q=Path(DISABLED_POWER_QUERY_FOLDER).expanduser().resolve() if clean(DISABLED_POWER_QUERY_FOLDER) else None
    return Config(q,Path(MAPPING_WORKBOOK).expanduser().resolve(),MAPPING_SHEET,
                  Path(TMDL_TABLES_INPUT_FOLDER).expanduser().resolve(),Path(TMDL_OUTPUT_FOLDER).expanduser().resolve(),
                  Path(LOG_OUTPUT_FOLDER).expanduser().resolve(),key(DIRECTION),DRY_RUN)

def validate(c:Config):
    if c.direction not in {"target","legacy"}: raise ValueError("DIRECTION must be target or legacy")
    if not c.workbook.is_file(): raise FileNotFoundError(c.workbook)
    if not c.tmdl_in.is_dir(): raise FileNotFoundError(c.tmdl_in)
    if c.pq_folder and not c.pq_folder.is_dir(): raise FileNotFoundError(c.pq_folder)
    if c.tmdl_in==c.tmdl_out or relative(c.tmdl_out,c.tmdl_in) or relative(c.tmdl_in,c.tmdl_out):
        raise ValueError("TMDL input/output folders must be separate and non-nested")
    if TMDL_OUTPUT_MODE not in {"tmdl_view_script","pbip_definition"}: raise ValueError("Invalid TMDL_OUTPUT_MODE")
    c.tmdl_out.mkdir(parents=True,exist_ok=True); c.log_out.mkdir(parents=True,exist_ok=True)

# =============================================================================
# TYPE AND MAPPING
# =============================================================================
def type_decision(ltable,lcol,lraw,traw,direction)->TypeDecision:
    ln,tn=norm_type(lraw),norm_type(traw)
    for o in COLUMN_TYPE_OVERRIDES:
        if key(o.get("legacy_table",""))==key(ltable) and key(o.get("legacy_column",""))==key(lcol):
            return TypeDecision(lraw,traw,ln,tn,norm_type(o["final_type"]),"Column override",clean(o.get("description")))
    if direction=="legacy": return TypeDecision(lraw,traw,ln,tn,ln,"Legacy destination type","",True)
    for r in LEGACY_TYPE_RETENTION_RULES:
        exact=key(r.get("match_mode","normalized"))=="exact"
        ok=(raw_type(r["legacy_type"])==raw_type(lraw) and raw_type(r["target_type"])==raw_type(traw)) if exact else (norm_type(r["legacy_type"])==ln and norm_type(r["target_type"])==tn)
        if ok: return TypeDecision(lraw,traw,ln,tn,ln,"Legacy retention rule",clean(r.get("description")),True)
    return TypeDecision(lraw,traw,ln,tn,tn,"Target type")

def load_mapping(c:Config)->ModelMap:
    df=pd.read_excel(c.workbook,sheet_name=c.sheet,engine="openpyxl",dtype=str); df.columns=[clean(x) for x in df.columns]
    miss=[x for x in REQ if x not in df.columns]
    if miss: raise ValueError(f"Missing mapping columns: {miss}")
    for x in REQ: df[x]=df[x].fillna("").map(clean)
    pairs={}; cols=[]; errors=[]; scoped={}
    for n,row in enumerate(df.to_dict("records"),2):
        lt,tt=row[REQ[0]],row[REQ[1]]
        if not lt or not tt: errors.append(f"Row {n}: both table names required"); continue
        split_table(lt); split_table(tt)
        src,dst=(lt,tt) if c.direction=="target" else (tt,lt); sk=key(src)
        if sk in pairs and key(pairs[sk][1])!=key(dst): errors.append(f"Row {n}: table maps to multiple destinations")
        pairs[sk]=(src,dst)
        lc,tc,lty,tty=row[REQ[2]],row[REQ[4]],row[REQ[3]],row[REQ[5]]
        if not lc and not tc: continue
        if not lc or not tc: continue  # target-only / legacy-only is valid but not paired
        if not lty or not tty: errors.append(f"Row {n}: paired columns require both types"); continue
        scope=(key(lt),key(lc)); dest=(key(tt),key(tc))
        if scope in scoped and scoped[scope]!=dest: errors.append(f"Row {n}: {lt}.{lc} maps to multiple targets"); continue
        scoped[scope]=dest
        d=type_decision(lt,lc,lty,tty,c.direction); cols.append(ColMap(lt,tt,lc,tc,lty,tty,d))
    if errors: raise ValueError("Mapping validation failed:\n- "+"\n- ".join(errors))
    by=defaultdict(list)
    for x in cols: by[key(x.src_table)].append(x)
    return ModelMap(pairs,dict(by),cols)

# =============================================================================
# LOGGING
# =============================================================================
SHEETS={
"1_Run Summary":["Item","Value"],
"2_Files Processed":["Run ID","Run Timestamp","Input File","Input Path","File Type","Classification","Eligible","Relevant Mapping Found","Changed","Output Existed Before Run","Output Action","Output File","Output Path","Write Status","Message","Input SHA-256","Previous Output SHA-256","Generated Output SHA-256"],
"3_Table Changes":["Run ID","Input File","File Type","Query or Model Table","Change Location","Old Table","New Table","SQL Alias","Alias Preserved","Replacement Count","Result"],
"4_Column Changes":["Run ID","Input File","Model Table","Model Column","Old Source Column","New Source Column","Changed In","Model Name Changed","Result"],
"5_Data Type Decisions":["Run ID","Input File","Table","Model Column","Legacy Type","Target Type","Final Type","Decision Source","Rule Description","SQL Cast Added","M Step Added","TMDL Updated"],
"6_Native SQL Changes":["Run ID","Input File","Query Name","Change Type","Old Value","New Value","Occurrences","Safe Replacement","Message"],
"7_M Query Changes":["Run ID","Input File","Query Name","M Step","Change Type","Old Value","New Value","Result"],
"8_Protected Content":["Run ID","Input File","Protected Type","Object Name","Count","Action","Reason"],
"9_Warnings":["Run ID","Input File","Warning Code","Category","Object","Warning","Impact","Recommended Review"],
"10_Errors":["Run ID","Input File","Stage","Error Code","Error Message","Existing Output Preserved","Processing Continued"]}

class Log:
    def __init__(self,rid):
        self.rid=rid
        self.timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.rows={sheet:[] for sheet in SHEETS}
    def add(self,s,**kw):
        r={c:"" for c in SHEETS[s]}; r.update(kw)
        if "Run ID" in r:r["Run ID"]=self.rid
        if "Run Timestamp" in r:r["Run Timestamp"]=self.timestamp
        self.rows[s].append(r)

def flush(ch:Changes,log:Log):
    for s,rows in [("3_Table Changes",ch.table),("4_Column Changes",ch.column),("5_Data Type Decisions",ch.dtype),("6_Native SQL Changes",ch.sql),("7_M Query Changes",ch.m),("8_Protected Content",ch.protected),("9_Warnings",ch.warnings),("10_Errors",ch.errors)]:
        for r in rows: log.add(s,**r)

def print_progress(kind:str, inp:Path, action:str, message:str):
    if LIVE_CONSOLE_PROGRESS:
        print(f"[{kind}] {inp.name}: {action} - {message}")


def write_result(inp:Path,out:Path,text:str,nl:str,ch:Changes,kind:str,log:Log):
    formatted=newline(text,nl)
    exists=out.exists()
    input_text,_=read_text(inp)
    input_hash=sha(input_text)
    previous_hash=""
    if exists:
        try: previous_hash=sha(read_text(out)[0])
        except Exception: previous_hash="Unreadable"
    generated_hash=sha(formatted) if ch.changed else ""
    action="Skipped"; status="Not written"; msg="No applicable changes"
    if ch.changed:
        if exists:
            if previous_hash==generated_hash:
                action="Unchanged";status="Write skipped";msg="Existing output matches regenerated content"
            elif CFG.dry_run:
                action="Would overwrite";status="Dry run";msg="Existing output differs; dry run prevented overwrite"
            elif OVERWRITE_EXISTING_OUTPUTS:
                atomic_write(out,formatted);action="Overwritten";status="Success";msg="Existing output found and overwritten"
            else:
                action="Failed";status="Not written";msg="Output exists and overwrite is disabled"
        elif CFG.dry_run:
            action="Would create";status="Dry run";msg="Dry run prevented creation"
        else:
            atomic_write(out,formatted);action="Created";status="Success";msg="New migrated output created"
    elif exists:
        action="Stale output retained";status="Not written";msg="Existing output found, but current input produced no applicable changes"
        if DELETE_STALE_OUTPUTS and not CFG.dry_run:
            out.unlink();action="Stale output deleted";status="Success";msg="Existing stale output deleted"
    print_progress(kind,inp,action,msg)
    log.add("2_Files Processed",**{
        "Input File":inp.name,"Input Path":str(inp),"File Type":kind,"Classification":ch.classification,
        "Eligible":"Yes","Relevant Mapping Found":"Yes" if ch.relevant else "No","Changed":"Yes" if ch.changed else "No",
        "Output Existed Before Run":"Yes" if exists else "No","Output Action":action,
        "Output File":out.name if action!="Skipped" else "","Output Path":str(out) if action!="Skipped" else "",
        "Write Status":status,"Message":msg,"Input SHA-256":input_hash,
        "Previous Output SHA-256":previous_hash,"Generated Output SHA-256":generated_hash})

# =============================================================================
# POWER QUERY AND SQL PARSING
# =============================================================================
def decode_m(s): return re.sub(r"#\(lf\)","\n",re.sub(r"#\(tab\)","\t",re.sub(r"#\(cr\)","\r",s,flags=re.I),flags=re.I),flags=re.I).replace('""','"')
def encode_m(s): return s.replace('"','""').replace("\r\n","\n").replace("\r","#(cr)").replace("\t","#(tab)").replace("\n","#(lf)")
def extract_query(m):
    q=re.search(r"\[\s*Query\s*=\s*\"",m,re.I)
    if not q:return None
    i=q.end(); j=i
    while j<len(m):
        if m[j]=='"':
            if j+1<len(m) and m[j+1]=='"': j+=2; continue
            return i,j,decode_m(m[i:j])
        j+=1
    raise ValueError("Unclosed Query string")
def classify_m(m):
    if re.search(r'\[\s*Query\s*=\s*"|Value\.NativeQuery',m,re.I): return "Native SQL"
    if "Table.NestedJoin" in m:return "Non-native join"
    return "Non-native"

def replace_tables(text,mapx,ch,input_file,location):
    out=text
    for src,dst in sorted(mapx.table_pairs.values(),key=lambda x:len(x[0]),reverse=True):
        ss,so=split_table(src); ds,do=split_table(dst); count=0
        for pat,rep in [(rf"(?<!\w){re.escape(ss)}\.{re.escape(so)}(?!\w)",dst),(rf"\[{re.escape(ss)}\]\.\[{re.escape(so)}\]",f"[{ds}].[{do}]")]:
            out,n=re.subn(pat,rep,out,flags=re.I); count+=n
        if count: ch.changed=ch.relevant=True; ch.table.append({"Input File":input_file,"Location":location,"Old Table":src,"New Table":dst,"Count":count})
    return out

def _require_sqlglot():
    if parse_one is None or build_scope is None or exp is None:
        raise ImportError(
            "Native SQL migration requires SQLGlot. Install it in the Python "
            "environment used to run this file with: python -m pip install sqlglot"
        )


def _canonical_identifier(value):
    return re.sub(r"[^A-Za-z0-9]+", "", clean(value)).casefold()


def _table_parts(table):
    parts=[]
    catalog=table.args.get("catalog")
    database=table.args.get("db")
    if catalog:
        parts.append(catalog.name)
    if database:
        parts.append(database.name)
    parts.append(table.name)
    return ".".join(parts)


def _mapping_pair_for_sql_table(mapx, table_name):
    direct=mapx.table_pairs.get(key(table_name))
    if direct:
        return direct
    matches=[
        (source,target)
        for source,target in mapx.table_pairs.values()
        if key(target)==key(table_name)
    ]
    return matches[0] if len(matches)==1 else None


def _column_mapping_for_table(mapx, source_table, column_name):
    mappings=mapx.cols_by_src_table.get(key(source_table),[])
    source_matches=[cm for cm in mappings if key(cm.src_col)==key(column_name)]
    source_targets={key(cm.dst_col):cm for cm in source_matches}
    if len(source_targets)==1:
        return next(iter(source_targets.values())),"legacy"
    target_matches=[cm for cm in mappings if key(cm.dst_col)==key(column_name)]
    target_targets={key(cm.dst_col):cm for cm in target_matches}
    if len(target_targets)==1:
        return next(iter(target_targets.values())),"target"
    return None,""


def _scope_output_names(source_scope):
    expression=source_scope.expression
    if not isinstance(expression,exp.Select):
        return set()
    names=set()
    for projection in expression.expressions:
        name=projection.alias_or_name
        if name:
            names.add(key(name))
    return names


def _physical_sources(scope,mapx):
    result={}
    for alias,source in scope.sources.items():
        if not isinstance(source,exp.Table):
            continue
        current=_table_parts(source)
        pair=_mapping_pair_for_sql_table(mapx,current)
        if pair:
            result[key(alias)]={
                "alias":alias,
                "table":source,
                "source_table":pair[0],
                "target_table":pair[1],
            }
    return result


def _derived_sources(scope):
    result={}
    for alias,source in scope.sources.items():
        if hasattr(source,"expression") and not isinstance(source,exp.Table):
            result[key(alias)]={
                "alias":alias,
                "scope":source,
                "outputs":_scope_output_names(source),
            }
    return result


def _resolve_sql_column(column,scope,mapx,physical,derived):
    """Resolve a SQL column without crossing a query-scope boundary."""
    if column.table:
        alias_key=key(column.table)
        if alias_key in derived:
            return {"kind":"derived","source":derived[alias_key]}
        item=physical.get(alias_key)
        if not item:
            # Correlated outer references are valid. Unknown local aliases are
            # checked separately by _validate_alias_scope.
            return None
        mapping,status=_column_mapping_for_table(
            mapx,item["source_table"],column.name
        )
        if mapping:
            return {
                "kind":"physical","source":item,
                "mapping":mapping,"status":status,
            }
        return None

    # Prefer an exact output supplied by one derived table. This prevents a
    # stable derived alias such as Percentage from being mistaken for a physical
    # target column in another table visible in the same SELECT.
    derived_matches=[
        item for item in derived.values()
        if key(column.name) in item["outputs"]
    ]
    if len(derived_matches)==1:
        return {"kind":"derived","source":derived_matches[0]}
    if len(derived_matches)>1:
        raise ValueError(
            "AMBIGUOUS_DERIVED_SQL_COLUMN: "
            f"{column.name} is exposed by multiple derived sources: "
            +", ".join(item["alias"] for item in derived_matches)
        )

    candidates=[]
    for item in physical.values():
        mapping,status=_column_mapping_for_table(
            mapx,item["source_table"],column.name
        )
        if mapping:
            candidates.append((item,mapping,status))
    unique={
        (key(item["alias"]),key(mapping.dst_col)):(item,mapping,status)
        for item,mapping,status in candidates
    }
    if len(unique)==1:
        item,mapping,status=next(iter(unique.values()))
        return {
            "kind":"physical","source":item,
            "mapping":mapping,"status":status,
        }
    if len(unique)>1:
        raise ValueError(
            "AMBIGUOUS_SQL_COLUMN: Unqualified column "
            f"{column.name!r} matches multiple physical sources: "
            +", ".join(
                f"{item['alias']}:{item['source_table']}"
                for item,mapping,status in unique.values()
            )
        )
    return None


def _target_sql_column(source,mapping):
    return exp.column(mapping.dst_col,table=source["alias"],quoted=True)


def _destination_physical_type(mapping):
    return (
        mapping.decision.target_norm
        if CFG.direction=="target"
        else mapping.decision.legacy_norm
    )


def _cast_projection_if_required(expression,mapping):
    if mapping.decision.final_norm==_destination_physical_type(mapping):
        return expression,False
    sql_type=SQL_TYPES.get(mapping.decision.final_norm)
    if not sql_type:
        return expression,False
    return exp.Cast(
        this=expression,
        to=exp.DataType.build(sql_type,dialect="tsql"),
    ),True


def _parent_scope_aliases(scope):
    aliases=set()
    parent=scope.parent
    while parent is not None:
        aliases.update(key(alias) for alias in parent.sources)
        parent=parent.parent
    return aliases


def _validate_alias_scope(scope):
    local={key(alias) for alias in scope.sources}
    outer=_parent_scope_aliases(scope)
    for column in scope.columns:
        if column.table and key(column.table) not in local|outer:
            raise ValueError(
                "SQL_ALIAS_NOT_IN_SCOPE: "
                f"Alias {column.table!r} is referenced inside a scope where it "
                "is not defined. Column: "
                f"{column.sql(dialect='tsql')}"
            )


def _migrate_select_projection(projection,scope,mapx,physical,derived,ch,input_file):
    existing_alias=projection.alias
    inner=projection.this if isinstance(projection,exp.Alias) else projection
    if not isinstance(inner,exp.Column):
        return projection
    resolved=_resolve_sql_column(inner,scope,mapx,physical,derived)
    if not resolved or resolved["kind"]!="physical":
        return projection
    mapping=resolved["mapping"]
    source=resolved["source"]
    output_name=existing_alias or inner.name
    migrated=_target_sql_column(source,mapping)
    migrated,cast_added=_cast_projection_if_required(migrated,mapping)
    result=exp.alias_(migrated,output_name,quoted=True)
    if result.sql(dialect="tsql")!=projection.sql(dialect="tsql"):
        ch.changed=ch.relevant=True
        ch.sql.append({
            "Input File":input_file,"Query Name":"",
            "Change Type":"Scoped physical projection",
            "Old Value":projection.sql(dialect="tsql"),
            "New Value":result.sql(dialect="tsql"),
            "Occurrences":1,"Safe Replacement":"Yes",
            "Message":(
                f"{source['source_table']} -> {source['target_table']}; "
                f"output alias {output_name!r} preserved"
            ),
        })
        ch.dtype.append({
            "Input File":input_file,"Table":mapping.legacy_table,
            "Model Column":output_name,"Legacy Type":mapping.legacy_type,
            "Target Type":mapping.target_type,
            "Final Type":mapping.decision.final_norm,
            "Decision Source":mapping.decision.source,
            "Rule Description":mapping.decision.description,
            "SQL Cast Added":"Yes" if cast_added else "No",
            "M Step Added":"No","TMDL Updated":"No",
        })
    return result


def _is_direct_select_projection(column):
    parent=column.parent
    while isinstance(parent,exp.Paren):
        parent=parent.parent
    if isinstance(parent,exp.Select):
        return True
    return (
        isinstance(parent,exp.Alias)
        and parent.this is column
        and isinstance(parent.parent,exp.Select)
    )


def _migrate_scope_columns(scope,mapx,physical,derived,ch,input_file):
    for column in list(scope.columns):
        if _is_direct_select_projection(column):
            continue
        resolved=_resolve_sql_column(column,scope,mapx,physical,derived)
        if not resolved:
            continue
        if resolved["kind"]=="derived":
            if not column.table:
                alias=resolved["source"]["alias"]
                replacement=exp.column(column.name,table=alias,quoted=True)
                old=column.sql(dialect="tsql")
                column.replace(replacement)
                ch.changed=ch.relevant=True
                ch.sql.append({
                    "Input File":input_file,"Query Name":"",
                    "Change Type":"Qualified derived output",
                    "Old Value":old,"New Value":replacement.sql(dialect="tsql"),
                    "Occurrences":1,"Safe Replacement":"Yes",
                    "Message":"Resolved from derived-table output schema",
                })
            continue
        mapping=resolved["mapping"]
        if resolved["status"]=="target":
            continue
        replacement=_target_sql_column(resolved["source"],mapping)
        old=column.sql(dialect="tsql")
        column.replace(replacement)
        ch.changed=ch.relevant=True
        ch.sql.append({
            "Input File":input_file,"Query Name":"",
            "Change Type":"Scoped physical column",
            "Old Value":old,"New Value":replacement.sql(dialect="tsql"),
            "Occurrences":1,"Safe Replacement":"Yes",
            "Message":f"Resolved within {resolved['source']['source_table']}",
        })


def _replace_sql_tables(expression,mapx,ch,input_file):
    """Replace physical tables while preserving every valid qualifier.

    A table without an explicit alias is implicitly referenced by its object
    name. When the object name changes, retain that old implicit qualifier as an
    explicit alias. Existing explicit aliases are left unchanged.
    """
    for table in list(expression.find_all(exp.Table)):
        current=_table_parts(table)
        pair=_mapping_pair_for_sql_table(mapx,current)
        if not pair or key(current)==key(pair[1]):
            continue

        schema,name=split_table(pair[1])
        old=current
        old_object_name=table.name
        explicit_alias=table.args.get("alias")

        # If the legacy SQL used the table's object name as a qualifier, that
        # qualifier must remain visible after the physical object is renamed.
        if explicit_alias is None:
            table.set(
                "alias",
                exp.TableAlias(
                    this=exp.to_identifier(old_object_name,quoted=True)
                ),
            )
            preserved_alias=old_object_name
            alias_origin="Implicit object-name qualifier preserved explicitly"
        else:
            preserved_alias=table.alias
            alias_origin="Existing explicit alias preserved"

        table.set("catalog",None)
        table.set("db",exp.to_identifier(schema,quoted=True))
        table.set("this",exp.to_identifier(name,quoted=True))

        ch.changed=ch.relevant=True
        ch.table.append({
            "Input File":input_file,
            "File Type":"Native SQL",
            "Query or Model Table":"",
            "Change Location":"SQL AST",
            "Old Table":old,
            "New Table":pair[1],
            "SQL Alias":preserved_alias,
            "Alias Preserved":"Yes",
            "Replacement Count":1,
            "Result":alias_origin,
        })


def migrate_native_sql(sql,mapx,ch,input_file):
    """Generic mapping-driven native SQL migration with scope lineage."""
    _require_sqlglot()
    try:
        expression=parse_one(sql,read="tsql")
    except ParseError as exc:
        raise ValueError(f"NATIVE_SQL_PARSE_FAILED: {exc}") from exc

    root=build_scope(expression)
    if root is None:
        raise ValueError("NATIVE_SQL_SCOPE_FAILED: No SQL scope was built")

    scopes=list(root.traverse())
    for scope in reversed(scopes):
        _validate_alias_scope(scope)
        physical=_physical_sources(scope,mapx)
        derived=_derived_sources(scope)
        if isinstance(scope.expression,exp.Select):
            scope.expression.set(
                "expressions",
                [
                    _migrate_select_projection(
                        projection,scope,mapx,physical,derived,ch,input_file
                    )
                    for projection in scope.expression.expressions
                ],
            )
        _migrate_scope_columns(
            scope,mapx,physical,derived,ch,input_file
        )

    _replace_sql_tables(expression,mapx,ch,input_file)
    migrated=expression.sql(dialect="tsql",pretty=True)

    try:
        reparsed=parse_one(migrated,read="tsql")
        rebuilt=build_scope(reparsed)
        if rebuilt is None:
            raise ValueError("Generated SQL has no scope tree")
        for scope in rebuilt.traverse():
            _validate_alias_scope(scope)
    except Exception as exc:
        raise ValueError(f"GENERATED_SQL_VALIDATION_FAILED: {exc}") from exc
    return migrated


def protect_odbc_scalar_escapes(sql):
    """Protect remaining `{fn ...}` expressions with stable SQL literals.

    Known functions are normalized before this runs. Remaining ODBC escapes are
    opaque driver syntax and must not be parsed as tables, columns, or functions.
    """
    protected={}
    output=[]
    index=0
    token_number=0
    length=len(sql)

    while index < length:
        match=re.match(r'\{\s*fn\b',sql[index:],re.IGNORECASE)
        if sql[index]!='{' or not match:
            output.append(sql[index]);index+=1;continue

        start=index
        depth=0
        in_single=in_double=in_bracket=False
        while index < length:
            char=sql[index]
            if in_single:
                if char=="'":
                    if index+1 < length and sql[index+1]=="'":
                        index+=2;continue
                    in_single=False
            elif in_double:
                if char=='"':
                    if index+1 < length and sql[index+1]=='"':
                        index+=2;continue
                    in_double=False
            elif in_bracket:
                if char==']':
                    if index+1 < length and sql[index+1]==']':
                        index+=2;continue
                    in_bracket=False
            else:
                if char=="'":in_single=True
                elif char=='"':in_double=True
                elif char=='[':in_bracket=True
                elif char=='{':depth+=1
                elif char=='}':
                    depth-=1
                    if depth==0:
                        index+=1;break
            index+=1

        if depth!=0:
            raise ValueError(
                "ODBC_ESCAPE_PARSE_FAILED: Unclosed ODBC scalar escape at "
                f"SQL position {start}"
            )
        original=sql[start:index]
        token_name=f"__PBI_ODBC_SCALAR_ESCAPE_{token_number:04d}__"
        if token_name in sql:
            raise ValueError(f"ODBC_ESCAPE_TOKEN_COLLISION: {token_name}")
        literal="'"+token_name+"'"
        protected[token_name]=original
        output.append(literal)
        token_number+=1

    return ''.join(output),protected


def restore_odbc_scalar_escapes(sql,protected):
    """Restore each protected escape exactly once after SQL migration."""
    result=sql
    for token_name,original in protected.items():
        pattern=re.compile(r"(?:N)?'"+re.escape(token_name)+r"'",re.IGNORECASE)
        matches=list(pattern.finditer(result))
        if len(matches)!=1:
            raise ValueError(
                "ODBC_ESCAPE_RESTORE_FAILED: Protection token "
                f"{token_name} occurred {len(matches)} times after SQL migration; "
                "expected 1"
            )
        result=pattern.sub(lambda _m:original,result,count=1)
    leftovers=re.findall(
        r"(?:N)?'__PBI_ODBC_SCALAR_ESCAPE_\d{4}__'",result,re.IGNORECASE
    )
    if leftovers:
        raise ValueError(
            "ODBC_ESCAPE_RESTORE_FAILED: Unrestored tokens: "+", ".join(leftovers)
        )
    return result


def normalize_target_sql_functions(sql,ch,input_file):
    """Convert configured source SQL functions to Azure Synapse T-SQL.

    The scanner skips strings, identifiers, and comments. Both bare functions
    and ODBC `{fn ...}` wrappers are supported. Every conversion is logged before
    later SQL processing, and the log is flushed even if a later stage fails.
    """
    function_map={
        "curdate":"CAST(GETDATE() AS date)",
    }
    output=[]
    changes=[]
    index=0
    length=len(sql)

    def consume_quoted(start,quote,end_quote=None):
        end_quote=end_quote or quote
        cursor=start+1
        while cursor < length:
            if sql[cursor]==end_quote:
                if cursor+1 < length and sql[cursor+1]==end_quote:
                    cursor+=2;continue
                return cursor+1
            cursor+=1
        raise ValueError(f"SQL_LEXER_UNCLOSED_{ord(quote)}")

    while index < length:
        char=sql[index]
        if char=="'":
            end=consume_quoted(index,"'");output.append(sql[index:end]);index=end;continue
        if char=='"':
            end=consume_quoted(index,'"');output.append(sql[index:end]);index=end;continue
        if char=='[':
            end=consume_quoted(index,'[',']');output.append(sql[index:end]);index=end;continue
        if sql.startswith('--',index):
            end=sql.find('\n',index);end=length if end<0 else end
            output.append(sql[index:end]);index=end;continue
        if sql.startswith('/*',index):
            end=sql.find('*/',index+2)
            if end<0:raise ValueError("SQL_LEXER_UNCLOSED_BLOCK_COMMENT")
            end+=2;output.append(sql[index:end]);index=end;continue

        escape=re.match(
            r'\{\s*fn\s+(?P<name>[A-Za-z_][A-Za-z0-9_]*)\s*\(\s*\)\s*\}',
            sql[index:],re.IGNORECASE,
        )
        if escape and escape.group('name').casefold() in function_map:
            original=escape.group(0);target=function_map[escape.group('name').casefold()]
            output.append(target);changes.append((original,target,"ODBC scalar escape"))
            index+=escape.end();continue

        bare=re.match(
            r'(?P<name>[A-Za-z_][A-Za-z0-9_]*)\s*\(\s*\)',
            sql[index:],re.IGNORECASE,
        )
        if bare:
            name=bare.group('name').casefold()
            previous=sql[index-1] if index else ''
            if name in function_map and not (previous.isalnum() or previous in '_.'):
                original=bare.group(0);target=function_map[name]
                output.append(target);changes.append((original,target,"Bare scalar function"))
                index+=bare.end();continue

        output.append(char);index+=1

    for original,target,kind in changes:
        ch.sql.append({
            "Input File":input_file,"Query Name":"",
            "Change Type":"Target SQL function normalization",
            "Old Value":original,"New Value":target,"Occurrences":1,
            "Safe Replacement":"Yes",
            "Message":kind+" converted before SQL parsing",
        })
    if changes:ch.changed=ch.relevant=True
    return ''.join(output)


def find_odbc_scalar_escapes(sql):
    """Return executable `{fn ...}` expressions only.

    SQL comments, string literals, quoted identifiers, and bracket identifiers
    are skipped. This is essential for legacy queries that retain disabled
    CURDATE filters as comments.
    """
    found=[]
    index=0
    length=len(sql)

    def consume_delimited(start,closing):
        cursor=start+1
        while cursor<length:
            if sql[cursor]==closing:
                if cursor+1<length and sql[cursor+1]==closing:
                    cursor+=2;continue
                return cursor+1
            cursor+=1
        raise ValueError("SQL_LEXER_UNCLOSED_DELIMITED_TEXT")

    while index<length:
        if sql[index]=="'":
            index=consume_delimited(index,"'");continue
        if sql[index]=='"':
            index=consume_delimited(index,'"');continue
        if sql[index]=='[':
            index=consume_delimited(index,']');continue
        if sql.startswith('--',index):
            newline=sql.find('\n',index+2)
            index=length if newline<0 else newline+1
            continue
        if sql.startswith('/*',index):
            close=sql.find('*/',index+2)
            if close<0:raise ValueError("SQL_LEXER_UNCLOSED_BLOCK_COMMENT")
            index=close+2;continue

        if sql[index]!='{' or not re.match(r'\{\s*fn\b',sql[index:],re.IGNORECASE):
            index+=1;continue

        start=index
        depth=0
        in_single=in_double=in_bracket=False
        while index<length:
            char=sql[index]
            if in_single:
                if char=="'":
                    if index+1<length and sql[index+1]=="'":index+=2;continue
                    in_single=False
            elif in_double:
                if char=='"':
                    if index+1<length and sql[index+1]=='"':index+=2;continue
                    in_double=False
            elif in_bracket:
                if char==']':
                    if index+1<length and sql[index+1]==']':index+=2;continue
                    in_bracket=False
            else:
                if char=="'":in_single=True
                elif char=='"':in_double=True
                elif char=='[':in_bracket=True
                elif char=='{':depth+=1
                elif char=='}':
                    depth-=1
                    if depth==0:
                        index+=1;break
            index+=1
        if depth!=0:
            raise ValueError(
                f"ODBC_ESCAPE_PARSE_FAILED: unclosed escape at SQL position {start}"
            )
        found.append(sql[start:index])
    return found



def validate_no_unsupported_odbc_escapes(sql,ch,input_file):
    """Stop before AST migration when an unconfigured ODBC escape remains."""
    escapes=find_odbc_scalar_escapes(sql)
    if not escapes:return
    for expression in escapes:
        ch.warnings.append({
            "Input File":input_file,
            "Warning Code":"UNSUPPORTED_ODBC_SCALAR_ESCAPE",
            "Category":"Native SQL",
            "Object":expression,
            "Warning":"No target SQL normalization is configured for this ODBC escape",
            "Impact":"Native SQL migration was stopped before parsing",
            "Recommended Review":"Add an explicit source-function mapping or rewrite the source SQL",
        })
    raise ValueError(
        "UNSUPPORTED_ODBC_SCALAR_ESCAPE: Remaining ODBC expressions after "
        "function normalization: "+"; ".join(escapes)
    )


def process_native_m(m,mapx,ch,input_file):
    query=extract_query(m)
    if not query:
        return replace_tables(m,mapx,ch,input_file,"Value.NativeQuery")

    start,end,sql=query
    sql=normalize_target_sql_functions(sql,ch,input_file)
    validate_no_unsupported_odbc_escapes(sql,ch,input_file)
    migrated=migrate_native_sql(sql,mapx,ch,input_file)
    return m[:start]+encode_m(migrated)+m[end:]


def parse_list(s): return [x.replace('""','"') for x in re.findall(r'"((?:[^"]|"")*)"',s)]
def query_ref(v): return v.strip()[2:-1].replace('""','"') if v.strip().startswith('#"') else v.strip()
def discover_pq(c): return [] if not c.pq_folder else sorted([p for p in c.pq_folder.glob("*.txt") if not p.stem.casefold().endswith(GENERATED_SUFFIXES)],key=lambda p:p.name.casefold())
def infer_tables(text,mapx): return {k for k,(s,d) in mapx.table_pairs.items() if re.search(re.escape(s),text,re.I) or (lambda z:re.search(rf'Schema\s*=\s*"{re.escape(z[0])}"\s*,\s*Item\s*=\s*"{re.escape(z[1])}"',text,re.I))(split_table(s))}
def build_context(files,mapx):
    txt={key(p.stem):read_text(p)[0] for p in files}; ctx={k:infer_tables(v,mapx) for k,v in txt.items()}; changed=True
    while changed:
        changed=False
        for q,t in txt.items():
            for other in txt:
                if other!=q and (re.search(rf'#"{re.escape(other)}"',t,re.I) or re.search(rf'(?<!\w){re.escape(other)}(?!\w)',t,re.I)):
                    n=len(ctx[q]);ctx[q]|=ctx[other];changed|=len(ctx[q])!=n
    return ctx

def primary_tables(m):
    x=re.search(r'Table\.NestedJoin\(\s*(?P<q>#"(?:[^"]|"")*"|[A-Za-z_][\w ]*)\s*,',m,re.I|re.S)
    if x:return QUERY_CONTEXT.get(key(query_ref(x.group("q"))),set())
    x=re.search(r'(?m)^\s*Source\s*=\s*(?P<q>#"(?:[^"]|"")*"|[A-Za-z_][\w ]*)\s*(?:,|$)',m,re.I)
    return QUERY_CONTEXT.get(key(query_ref(x.group("q"))),set()) if x else set()
def replace_navigation(m,mapx,ch,input_file):
    tables=set()
    def f(x):
        src=f"{x.group('s')}.{x.group('i')}"; pair=mapx.table_pairs.get(key(src))
        if not pair:return x.group(0)
        ds,di=split_table(pair[1]);tables.add(key(pair[0]));ch.changed=ch.relevant=True;return f'[Schema="{ds}", Item="{di}"]'
    return re.sub(r'\[\s*Schema\s*=\s*"(?P<s>[^"]+)"\s*,\s*Item\s*=\s*"(?P<i>[^"]+)"\s*\]',f,m,flags=re.I),tables
def replace_m_fields(m,maps,ch,input_file):
    """Replace table-scoped M field references and source-column arguments.

    `maps` is already restricted to the resolved physical source table. This
    function changes source references, not semantic model names or generated M
    output aliases.
    """
    out=m

    def find_call_end(source, open_paren_index):
        depth=0
        in_string=False
        index=open_paren_index
        while index < len(source):
            character=source[index]
            if in_string:
                if character=='"':
                    if index+1 < len(source) and source[index+1]=='"':
                        index+=2
                        continue
                    in_string=False
            else:
                if character=='"':
                    in_string=True
                elif character=='(':
                    depth+=1
                elif character==')':
                    depth-=1
                    if depth==0:
                        return index+1
            index+=1
        return None

    # These functions use quoted strings as input column references. AddColumn
    # is intentionally excluded because its quoted string defines a new output.
    source_column_functions=(
        "RemoveColumns",
        "UnpivotOtherColumns",
        "SelectColumns",
        "ReorderColumns",
        "TransformColumnTypes",
        "TransformColumns",
        "Distinct",
        "Sort",
        "Group",
    )

    for cm in maps:
        if key(cm.src_col)==key(cm.dst_col):
            continue

        # Row-context references such as [Location]. Do not touch the TMDL
        # declaration name; this operates only inside the M source expression.
        row_pattern=r'\['+re.escape(cm.src_col)+r'\]'
        out,row_count=re.subn(
            row_pattern,
            '['+cm.dst_col+']',
            out,
            flags=re.IGNORECASE,
        )

        quoted_count=0
        for function_name in source_column_functions:
            call_pattern=re.compile(
                r'Table\.'+re.escape(function_name)+r'\(',
                re.IGNORECASE,
            )
            search_from=0
            while True:
                call=call_pattern.search(out,search_from)
                if not call:
                    break
                call_end=find_call_end(out,call.end()-1)
                if call_end is None:
                    search_from=call.end()
                    continue
                block=out[call.start():call_end]
                token_pattern=r'"'+re.escape(cm.src_col)+r'"'
                changed_block,count=re.subn(
                    token_pattern,
                    '"'+cm.dst_col+'"',
                    block,
                    flags=re.IGNORECASE,
                )
                if count:
                    out=out[:call.start()]+changed_block+out[call_end:]
                    call_end=call.start()+len(changed_block)
                    quoted_count+=count
                search_from=call_end

        # RenameColumns pairs are {"existing input", "new output"}. Replace
        # only the first item so deliberate output aliases remain unchanged.
        rename_pattern=re.compile(r'Table\.RenameColumns\(',re.IGNORECASE)
        search_from=0
        rename_count=0
        while True:
            call=rename_pattern.search(out,search_from)
            if not call:
                break
            call_end=find_call_end(out,call.end()-1)
            if call_end is None:
                search_from=call.end()
                continue
            block=out[call.start():call_end]
            pair_pattern=(
                r'\{\s*"'+re.escape(cm.src_col)
                +r'"\s*,\s*"(?P<output>(?:[^"]|"")*)"\s*\}'
            )
            changed_block,count=re.subn(
                pair_pattern,
                lambda match: '{"'+cm.dst_col+'", "'+match.group("output")+'"}',
                block,
                flags=re.IGNORECASE,
            )
            if count:
                out=out[:call.start()]+changed_block+out[call_end:]
                call_end=call.start()+len(changed_block)
                rename_count+=count
            search_from=call_end

        total=row_count+quoted_count+rename_count
        if total:
            ch.changed=True
            ch.relevant=True
            locations=[]
            if row_count:
                locations.append("M row expression")
            if quoted_count:
                locations.append("M quoted source-column argument")
            if rename_count:
                locations.append("M RenameColumns input")
            ch.column.append({
                "Input File":input_file,
                "Model Table":"",
                "Model Column":"",
                "Old Source Column":cm.src_col,
                "New Source Column":cm.dst_col,
                "Changed In":" and ".join(locations),
                "Model Name Changed":"No",
                "Result":f"Changed {total} occurrence(s)",
            })
            ch.m.append({
                "Input File":input_file,
                "Query Name":"",
                "M Step":"Column references",
                "Change Type":"Target source-column mapping",
                "Old Value":cm.src_col,
                "New Value":cm.dst_col,
                "Result":(
                    f"row={row_count}; quoted={quoted_count}; "
                    f"rename-input={rename_count}"
                ),
            })
    return out

def add_type_step(m,maps,ch,input_file):
    """Append a valid final Table.TransformColumnTypes step to an M let query.

    The new binding is inserted immediately before `in`. Only the expression
    returned by `in` is redirected to the new binding. Upstream bindings and
    Source references are never rewritten.
    """
    items=[]
    seen=set()
    for cm in maps:
        destination_type=(
            cm.decision.target_norm
            if CFG.direction=="target"
            else cm.decision.legacy_norm
        )
        output_column=cm.dst_col
        output_key=key(output_column)
        if (
            cm.decision.final_norm!=destination_type
            and output_key not in seen
            and cm.decision.final_norm in M_TYPES
        ):
            items.append((output_column,M_TYPES[cm.decision.final_norm],cm))
            seen.add(output_key)

    if not items or re.search(
        r'(?m)^\s*#"Changed Type Migration"\s*=',m
    ):
        return m

    # Restrict the operation to one complete M let expression.
    let_match=re.search(r'(?m)^\s*let\s*$',m)
    in_matches=list(re.finditer(r'(?m)^(?P<indent>\s*)in\s*$',m))
    if not let_match or not in_matches:
        ch.warnings.append({
            "Input File":input_file,
            "Warning Code":"M_TYPE_STEP_NO_LET_IN",
            "Category":"M query",
            "Object":"Changed Type Migration",
            "Warning":"A complete let/in expression was not found",
            "Impact":"The type-conversion step was not inserted",
            "Recommended Review":"Review the partition M expression",
        })
        return m

    in_match=in_matches[-1]
    in_end=in_match.end()
    returned_match=re.match(
        r'(?P<space>\s*)(?P<expr>#"(?:[^"]|"")*"|[A-Za-z_][A-Za-z0-9_ ]*)',
        m[in_end:],
    )
    if not returned_match:
        ch.warnings.append({
            "Input File":input_file,
            "Warning Code":"M_TYPE_STEP_RETURN_NOT_FOUND",
            "Category":"M query",
            "Object":"Changed Type Migration",
            "Warning":"The expression returned by in could not be identified",
            "Impact":"The type-conversion step was not inserted",
            "Recommended Review":"Review the final in expression",
        })
        return m

    previous_step=returned_match.group("expr")
    step_name='#"Changed Type Migration"'

    # Everything before `in` is the let-binding list. Ensure the prior final
    # binding ends with exactly one comma before appending the new binding.
    bindings=m[:in_match.start()].rstrip()
    if not bindings.endswith(','):
        bindings+=','

    in_indent=in_match.group("indent")
    binding_indent=in_indent + "    "
    detail_indent=binding_indent + "    "
    item_indent=detail_indent + "    "
    lines=(",\n").join(
        f'{item_indent}{{"{column.replace(chr(34), chr(34)*2)}", {m_type}}}'
        for column,m_type,_cm in items
    )

    step=(
        "\n"
        f"{binding_indent}{step_name} = Table.TransformColumnTypes(\n"
        f"{detail_indent}{previous_step},\n"
        f"{detail_indent}{{\n"
        f"{lines}\n"
        f"{detail_indent}}},\n"
        f'{detail_indent}[Culture="{M_CULTURE}", '
        "MissingField=MissingField.Ignore]\n"
        f"{binding_indent})\n"
    )

    # Preserve text after the returned expression, but redirect only that final
    # expression. This avoids changing Source or other upstream references.
    return_start=in_end+returned_match.start("expr")
    return_end=in_end+returned_match.end("expr")
    new_in_tail=m[in_match.start():return_start]+step_name+m[return_end:]
    result=bindings+step+new_in_tail

    # Lightweight structural validation catches the exact corruption previously
    # emitted: `let,`, a dangling `source ,`, or a binding before Source.
    if re.search(r'(?i)\blet\s*,',result):
        raise ValueError("M_TYPE_STEP_VALIDATION_FAILED: generated 'let,'")
    if result.count(step_name) < 2:
        raise ValueError(
            "M_TYPE_STEP_VALIDATION_FAILED: step binding/return linkage missing"
        )

    ch.changed=True
    ch.m.append({
        "Input File":input_file,
        "Query Name":"",
        "M Step":"Changed Type Migration",
        "Change Type":"Final step appended",
        "Old Value":previous_step,
        "New Value":step_name,
        "Result":f"{len(items)} conversion(s); final return redirected only",
    })
    return result

def _query_tables_from_reference(reference):
    """Resolve an exact M query reference through the disabled-query context."""
    return set(QUERY_CONTEXT.get(key(query_ref(reference)), set()))


def _mappings_for_query_reference(reference, mapx):
    mappings=[]
    table_keys=_query_tables_from_reference(reference)
    for table_key in table_keys:
        mappings.extend(mapx.cols_by_src_table.get(table_key, []))
    return table_keys,mappings


def _unique_target_column(column_name, mappings):
    legacy_matches=[cm for cm in mappings if key(cm.src_col)==key(column_name)]
    destinations={key(cm.dst_col):cm.dst_col for cm in legacy_matches}
    if len(destinations)==1:
        return next(iter(destinations.values()))
    # Rerun safety: a field already using the target name remains unchanged.
    if any(key(cm.dst_col)==key(column_name) for cm in mappings):
        return column_name
    return None


def _quoted_m_list(values):
    return ", ".join('"'+value.replace('"','""')+'"' for value in values)


def _replace_downstream_m_column(text, start_at, old_name, new_name):
    """Change references only after the expansion that creates the output."""
    if key(old_name)==key(new_name):
        return text,0
    before=text[:start_at]
    after=text[start_at:]
    count=0

    # Row-context references: [Date], [Description], and #["..."] form.
    after,n=re.subn(
        r'\['+re.escape(old_name)+r'\]',
        '['+new_name+']',
        after,
        flags=re.IGNORECASE,
    )
    count+=n
    after,n=re.subn(
        r'\[#"'+re.escape(old_name)+r'"\]',
        '[#"'+new_name+'"]',
        after,
        flags=re.IGNORECASE,
    )
    count+=n

    # Quoted column arguments in later M steps. This is intentionally limited
    # to text after the expansion; it does not modify query names or earlier
    # NestedJoin inputs.
    after,n=re.subn(
        r'"'+re.escape(old_name)+r'"',
        '"'+new_name+'"',
        after,
        flags=re.IGNORECASE,
    )
    count+=n
    return before+after,count


def migrate_joined_query_columns(m,mapx,ch,input_file):
    """Migrate Table.NestedJoin and ExpandTableColumn with branch lineage."""
    out=m
    nested_right_context={}

    join_pattern=re.compile(
        r'Table\.NestedJoin\(\s*'
        r'(?P<left>#"(?:[^"]|"")*"|[A-Za-z_][A-Za-z0-9_ ]*)\s*,\s*'
        r'\{(?P<leftkeys>.*?)\}\s*,\s*'
        r'(?P<right>#"(?:[^"]|"")*"|[A-Za-z_][A-Za-z0-9_ ]*)\s*,\s*'
        r'\{(?P<rightkeys>.*?)\}\s*,\s*'
        r'"(?P<nested>(?:[^"]|"")*)"',
        re.IGNORECASE|re.DOTALL,
    )

    search_from=0
    while True:
        match=join_pattern.search(out,search_from)
        if not match:
            break

        _,left_maps=_mappings_for_query_reference(match.group("left"),mapx)
        right_table_keys,right_maps=_mappings_for_query_reference(match.group("right"),mapx)
        nested_name=match.group("nested").replace('""','"')
        nested_right_context[key(nested_name)]={
            "query":query_ref(match.group("right")),
            "tables":right_table_keys,
            "mappings":right_maps,
        }

        old_left=parse_list(match.group("leftkeys"))
        old_right=parse_list(match.group("rightkeys"))
        new_left=[_unique_target_column(value,left_maps) or value for value in old_left]
        new_right=[_unique_target_column(value,right_maps) or value for value in old_right]

        block=match.group(0)
        block=block.replace(
            '{'+match.group("leftkeys")+'}',
            '{'+_quoted_m_list(new_left)+'}',
            1,
        )
        right_position=block.find(match.group("right"))
        prefix,suffix=block[:right_position],block[right_position:]
        suffix=suffix.replace(
            '{'+match.group("rightkeys")+'}',
            '{'+_quoted_m_list(new_right)+'}',
            1,
        )
        replacement=prefix+suffix
        out=out[:match.start()]+replacement+out[match.end():]
        search_from=match.start()+len(replacement)

        if new_left!=old_left or new_right!=old_right:
            ch.changed=ch.relevant=True
            ch.m.append({
                "Input File":input_file,
                "Query Name":"",
                "M Step":"Table.NestedJoin",
                "Change Type":"Branch-aware join keys",
                "Old Value":f"left={old_left}; right={old_right}",
                "New Value":f"left={new_left}; right={new_right}",
                "Result":f"Right query={query_ref(match.group('right'))}",
            })

    expand_pattern=re.compile(
        r'Table\.ExpandTableColumn\(\s*'
        r'(?P<table>[^,]+),\s*'
        r'"(?P<nested>(?:[^"]|"")*)"\s*,\s*'
        r'\{(?P<input>.*?)\}'
        r'(?:\s*,\s*\{(?P<output>.*?)\})?\s*\)',
        re.IGNORECASE|re.DOTALL,
    )

    search_from=0
    while True:
        match=expand_pattern.search(out,search_from)
        if not match:
            break

        nested_name=match.group("nested").replace('""','"')
        right_context=nested_right_context.get(key(nested_name))
        if not right_context:
            search_from=match.end()
            continue

        right_maps=right_context["mappings"]
        old_inputs=parse_list(match.group("input"))
        old_outputs=(
            parse_list(match.group("output"))
            if match.group("output") is not None
            else list(old_inputs)
        )

        if len(old_inputs)!=len(old_outputs):
            ch.warnings.append({
                "Input File":input_file,
                "Warning Code":"EXPAND_COLUMN_COUNT_MISMATCH",
                "Category":"M query",
                "Object":nested_name,
                "Warning":"Expand input and output column counts differ",
                "Impact":"Expand step was left unchanged",
                "Recommended Review":"Review Table.ExpandTableColumn manually",
            })
            search_from=match.end()
            continue

        new_inputs=[]
        new_outputs=[]
        rewires=[]
        for old_input,old_output in zip(old_inputs,old_outputs):
            target_input=_unique_target_column(old_input,right_maps) or old_input
            new_inputs.append(target_input)

            # The fourth list defines final M output names. Change it only when
            # it merely repeats the legacy inner column name. Deliberate aliases
            # such as "Mapping Date" remain unchanged.
            target_output=(
                target_input
                if key(old_output)==key(old_input)
                else old_output
            )
            new_outputs.append(target_output)
            if key(target_output)!=key(old_output):
                rewires.append((old_output,target_output))

        replacement=(
            'Table.ExpandTableColumn('
            +match.group("table").strip()+', '
            +'"'+nested_name.replace('"','""')+'", '
            +'{'+_quoted_m_list(new_inputs)+'}, '
            +'{'+_quoted_m_list(new_outputs)+'})'
        )
        out=out[:match.start()]+replacement+out[match.end():]
        downstream_start=match.start()+len(replacement)

        for old_name,new_name in rewires:
            out,count=_replace_downstream_m_column(
                out,downstream_start,old_name,new_name
            )
            if count:
                ch.column.append({
                    "Input File":input_file,
                    "Model Table":"",
                    "Model Column":"",
                    "Old Source Column":old_name,
                    "New Source Column":new_name,
                    "Changed In":"M downstream after joined expansion",
                    "Model Name Changed":"No",
                    "Result":f"Changed {count} downstream occurrence(s)",
                })

        if new_inputs!=old_inputs or new_outputs!=old_outputs:
            ch.changed=ch.relevant=True
            ch.m.append({
                "Input File":input_file,
                "Query Name":"",
                "M Step":"Table.ExpandTableColumn",
                "Change Type":"Right-query target columns",
                "Old Value":f"inputs={old_inputs}; outputs={old_outputs}",
                "New Value":f"inputs={new_inputs}; outputs={new_outputs}",
                "Result":f"Mapped using joined query {right_context['query']}",
            })

        search_from=downstream_start

    return out


def replace_unpivot_other_with_mapped_unpivot(m,mapx,table_context,ch,input_file):
    """Use an explicit mapped target-column list for unpivot operations."""
    out=m
    scoped_mappings=[]
    for table_key in table_context:
        scoped_mappings.extend(mapx.cols_by_src_table.get(table_key,[]))

    # Do not guess when the query resolves to zero or multiple physical tables.
    source_tables={key(cm.src_table) for cm in scoped_mappings}
    if len(source_tables)!=1:
        return out

    # Columns already removed earlier in this M query. Both legacy and target
    # names are accepted for rerun safety.
    removed_names=set()
    for removed_call in re.finditer(
        r'Table\.RemoveColumns\(.*?\{(?P<columns>.*?)\}\s*\)',
        out,
        re.IGNORECASE|re.DOTALL,
    ):
        removed_names.update(key(value) for value in parse_list(removed_call.group("columns")))

    pattern=re.compile(
        r'Table\.UnpivotOtherColumns\(\s*'
        r'(?P<table>[^,]+),\s*'
        r'\{(?P<keep>.*?)\}\s*,\s*'
        r'"(?P<attribute>(?:[^"]|"")*)"\s*,\s*'
        r'"(?P<value>(?:[^"]|"")*)"\s*\)',
        re.IGNORECASE|re.DOTALL,
    )

    search_from=0
    while True:
        match=pattern.search(out,search_from)
        if not match:
            break

        keep_names=parse_list(match.group("keep"))
        keep_keys={key(value) for value in keep_names}

        explicit=[]
        seen=set()
        for cm in scoped_mappings:
            source_key=key(cm.src_col)
            target_key=key(cm.dst_col)
            if source_key in keep_keys or target_key in keep_keys:
                continue
            if source_key in removed_names or target_key in removed_names:
                continue
            if target_key in seen:
                continue
            explicit.append(cm.dst_col)
            seen.add(target_key)

        if not explicit:
            ch.warnings.append({
                "Input File":input_file,
                "Warning Code":"UNPIVOT_TARGET_COLUMNS_UNRESOLVED",
                "Category":"M query",
                "Object":"Table.UnpivotOtherColumns",
                "Warning":"No explicit mapped target columns were available for unpivot",
                "Impact":"The original UnpivotOtherColumns step was preserved",
                "Recommended Review":"Verify paired column mappings for the source table",
            })
            search_from=match.end()
            continue

        explicit_text=", ".join(
            '"'+value.replace('"','""')+'"' for value in explicit
        )
        replacement=(
            'Table.Unpivot('
            +match.group("table").strip()+', '
            +'{'+explicit_text+'}, '
            +'"'+match.group("attribute")+'", '
            +'"'+match.group("value")+'")'
        )
        out=out[:match.start()]+replacement+out[match.end():]
        search_from=match.start()+len(replacement)
        ch.changed=ch.relevant=True
        ch.m.append({
            "Input File":input_file,
            "Query Name":"",
            "M Step":"Table.Unpivot",
            "Change Type":"Explicit mapped target columns",
            "Old Value":match.group(0),
            "New Value":replacement,
            "Result":(
                "Converted UnpivotOtherColumns to explicit mapped unpivot; "
                "target-only columns are excluded"
            ),
        })

    return out


def migrate_replacevalue_columns(m,maps,ch,input_file):
    """Rewrite only Table.ReplaceValue columnsToSearch using mapped targets."""
    out=m
    mapping_by_source={}
    target_names=set()
    for cm in maps:
        mapping_by_source.setdefault(key(cm.src_col),set()).add(cm.dst_col)
        target_names.add(key(cm.dst_col))

    def call_end(source,open_index):
        depth=0
        in_string=False
        index=open_index
        while index < len(source):
            char=source[index]
            if in_string:
                if char=='"':
                    if index+1 < len(source) and source[index+1]=='"':
                        index+=2
                        continue
                    in_string=False
            else:
                if char=='"':
                    in_string=True
                elif char=='(':
                    depth+=1
                elif char==')':
                    depth-=1
                    if depth==0:
                        return index
            index+=1
        return None

    def top_level_arguments(block,open_offset,close_offset):
        spans=[]
        start=open_offset+1
        depth=0
        brace=0
        bracket=0
        in_string=False
        index=start
        while index < close_offset:
            char=block[index]
            if in_string:
                if char=='"':
                    if index+1 < close_offset and block[index+1]=='"':
                        index+=2
                        continue
                    in_string=False
            else:
                if char=='"': in_string=True
                elif char=='(': depth+=1
                elif char==')': depth-=1
                elif char=='{': brace+=1
                elif char=='}': brace-=1
                elif char=='[': bracket+=1
                elif char==']': bracket-=1
                elif char==',' and depth==0 and brace==0 and bracket==0:
                    spans.append((start,index))
                    start=index+1
            index+=1
        spans.append((start,close_offset))
        return spans

    pattern=re.compile(r'Table\.ReplaceValue\s*\(',re.IGNORECASE)
    search_from=0
    while True:
        match=pattern.search(out,search_from)
        if not match:
            break
        open_index=out.find('(',match.start())
        close_index=call_end(out,open_index)
        if close_index is None:
            raise ValueError(
                "M_REPLACEVALUE_PARSE_FAILED: unmatched parenthesis"
            )
        argument_spans=top_level_arguments(out,open_index,close_index)
        if len(argument_spans)!=5:
            raise ValueError(
                "M_REPLACEVALUE_PARSE_FAILED: expected 5 arguments, found "
                f"{len(argument_spans)}"
            )

        list_start,list_end=argument_spans[4]
        column_list=out[list_start:list_end]
        changes=[]

        def replace_column(match_object):
            original=match_object.group(1).replace('""','"')
            candidates=mapping_by_source.get(key(original),set())
            if len(candidates)==1:
                target=next(iter(candidates))
                if key(target)!=key(original):
                    changes.append((original,target))
                return '"'+target.replace('"','""')+'"'
            if len(candidates)>1:
                raise ValueError(
                    "AMBIGUOUS_M_REPLACEVALUE_COLUMN: "
                    f"{original!r} maps to {sorted(candidates)}"
                )
            if key(original) in target_names:
                return match_object.group(0)
            return match_object.group(0)

        migrated_list=re.sub(
            r'"((?:[^"]|"")*)"',replace_column,column_list
        )
        if migrated_list!=column_list:
            out=out[:list_start]+migrated_list+out[list_end:]
            delta=len(migrated_list)-len(column_list)
            close_index+=delta
            ch.changed=ch.relevant=True
            for old_name,new_name in changes:
                ch.column.append({
                    "Input File":input_file,
                    "Model Table":"",
                    "Model Column":"",
                    "Old Source Column":old_name,
                    "New Source Column":new_name,
                    "Changed In":"Table.ReplaceValue columnsToSearch",
                    "Model Name Changed":"No",
                    "Result":"Changed 1 occurrence",
                })
            ch.m.append({
                "Input File":input_file,
                "Query Name":"",
                "M Step":"Table.ReplaceValue",
                "Change Type":"Mapped columnsToSearch",
                "Old Value":column_list.strip(),
                "New Value":migrated_list.strip(),
                "Result":f"{len(changes)} target column replacement(s)",
            })
        search_from=close_index+1
    return out


def process_non_native(m,mapx,ch,input_file):
    out,direct=replace_navigation(m,mapx,ch,input_file)
    out=replace_tables(out,mapx,ch,input_file,"M source")

    # Joined-query columns must be migrated before general primary-source column
    # replacement. ExpandTableColumn uses the right-side query's schema.
    out=migrate_joined_query_columns(out,mapx,ch,input_file)

    context=direct or primary_tables(out)
    maps=[]
    for table_key in context:
        maps+=mapx.cols_by_src_table.get(table_key,[])
    out=replace_m_fields(out,maps,ch,input_file)
    out=migrate_replacevalue_columns(out,maps,ch,input_file)

    # Use an explicit mapped target list for unpivot. This prevents target-only
    # indicator fields such as Y/N columns from flowing into numeric Value /
    # RowSortOrder output columns.
    out=replace_unpivot_other_with_mapped_unpivot(
        out,mapx,context,ch,input_file
    )

    out=add_type_step(out,maps,ch,input_file)
    return out

# =============================================================================
# TMDL SAFETY AND PROCESSING
# =============================================================================
def source_spans(text):
    lines=text.splitlines(True);offs=[];n=0
    for l in lines:offs.append(n);n+=len(l)
    out=[]
    for i,l in enumerate(lines):
        m=re.match(r'^(?P<z>\s*)source\s*=\s*$',l,re.I)
        if not m:continue
        base=len(m.group('z'));j=i+1
        while j<len(lines) and (not lines[j].strip() or len(lines[j])-len(lines[j].lstrip())>base):j+=1
        out.append((offs[i],offs[j] if j<len(offs) else len(text)))
    return out
def col_blocks(text):
    ms=list(re.finditer(r'(?m)^(?P<z>\s*)column\s+(?P<n>.+?)\s*$',text));out=[]
    for m in ms:
        z=len(m.group('z'));tail=text[m.end():];s=re.search(rf'(?m)^\s{{0,{z}}}(?:column|measure|hierarchy|partition|annotation)\b',tail);end=m.end()+s.start() if s else len(text);out.append((m.start(),end,strip_quotes(m.group('n'))))
    return out
def table_name(t):
    m=re.search(r'(?m)^table\s+(.+?)\s*$',t);return strip_quotes(m.group(1)) if m else ""
def physical_source_cols(t):return {key(strip_quotes(x)) for x in re.findall(r'(?m)^\s*sourceColumn:\s*(.+?)\s*$',t)}
def choose_mapping(old,candidates):
    a=[x for x in candidates if key(x.src_col)==key(old)]
    if len(a)==1:return a[0],"source"
    b=[x for x in candidates if key(x.dst_col)==key(old)]
    if len(b)==1:return b[0],"target"
    return (None,"")
def update_tmdl_cols(text,mapx,context,ch,input_file,native):
    """Update physical TMDL bindings without changing native-model types.

    Native SQL owns its result schema through SELECT aliases and CAST/CASE
    expressions. For native partitions, the semantic model's established
    dataType and sourceColumn are therefore preserved. For non-native M queries,
    sourceColumn and dataType can still be migrated to target outputs.
    """
    out=text
    candidates=[]
    for table_key in context:
        candidates+=mapx.cols_by_src_table.get(table_key,[])

    for st,en,modelcol in reversed(col_blocks(out)):
        block=out[st:en]
        source_match=re.search(
            r'(?m)^(?P<z>\s*)sourceColumn:\s*(?P<v>.+?)\s*$',
            block,
        )
        if not source_match:
            continue

        old_source=strip_quotes(source_match.group('v'))
        mapping,reason=choose_mapping(old_source,candidates)
        if not mapping:
            continue

        # Native SQL aliases target physical fields back to the existing output
        # contract. The TMDL sourceColumn must continue to bind to that final
        # SQL/M result name. A non-native Navigation query exposes target names
        # directly, so its sourceColumn can move to the target field.
        new_source=(
            old_source
            if native or reason=="target"
            else mapping.dst_col
        )

        new_block=(
            block[:source_match.start('v')]
            +quote_like(source_match.group('v'),new_source)
            +block[source_match.end('v'):]
        )

        type_changed=False
        old_type=""
        new_type=""
        type_match=re.search(
            r'(?m)^\s*dataType:\s*(?P<v>\S+)\s*$',
            new_block,
        )

        if type_match:
            old_type=type_match.group('v')

            if native:
                # Hard rule: native SQL tables keep their existing semantic
                # model type. Any physical conversion belongs in SQL, not in the
                # full-table TMDL replacement.
                new_type=old_type
            else:
                new_type=TMDL_TYPES[mapping.decision.final_norm]
                if key(old_type)!=key(new_type):
                    new_block=(
                        new_block[:type_match.start('v')]
                        +new_type
                        +new_block[type_match.end('v'):]
                    )
                    type_changed=True

        if new_block!=block:
            out=out[:st]+new_block+out[en:]
            ch.changed=ch.relevant=True
            ch.column.append({
                "Input File":input_file,
                "Model Table":table_name(text),
                "Model Column":modelcol,
                "Old Source Column":old_source,
                "New Source Column":new_source,
                "Changed In":"TMDL sourceColumn",
                "Model Name Changed":"No",
                "Result":(
                    "Native SQL output alias preserved; TMDL dataType preserved"
                    if native
                    else "Non-native target sourceColumn/type applied"
                ),
            })

        # Always make the native preservation decision visible in the log, even
        # when the TMDL text itself does not change.
        if native and type_match:
            ch.dtype.append({
                "Input File":input_file,
                "Table":mapping.legacy_table,
                "Model Column":modelcol,
                "Legacy Type":mapping.legacy_type,
                "Target Type":mapping.target_type,
                "Final Type":old_type,
                "Decision Source":"Preserve existing native TMDL dataType",
                "Rule Description":(
                    "Native SQL must return values compatible with the existing "
                    "semantic type; no TMDL dataType rewrite was performed"
                ),
                "SQL Cast Added":"No",
                "M Step Added":"No",
                "TMDL Updated":"No",
            })

    return out

def identity_blocks(text):
    out=[]
    for st,en,n in col_blocks(text):
        b=text[st:en];l=re.search(r'(?m)^\s*lineageTag:\s*(\S+)',b);m=re.match(r'(?P<z>\s*)column\s+.+?\s*$',text[st:text.find('\n',st) if text.find('\n',st)>=0 else en]);out.append((l.group(1) if l else "",n,st,text.find('\n',st) if text.find('\n',st)>=0 else en,m.group('z') if m else ""))
    return out
def restore_names(orig,mig,ch,input_file):
    result=mig;ot=re.search(r'(?m)^table\s+(?P<n>.+?)\s*$',orig);mt=re.search(r'(?m)^table\s+(?P<n>.+?)\s*$',result)
    if ot and mt and ot.group('n')!=mt.group('n'):result=result[:mt.start('n')]+ot.group('n')+result[mt.end('n'):]
    ob={l:(n) for l,n,st,en,z in identity_blocks(orig) if l}; reps=[]
    for l,n,st,en,z in identity_blocks(result):
        if l in ob and n!=ob[l]:reps.append((st,en,z+'column '+ob[l],n,ob[l]))
    for st,en,r,bad,good in sorted(reps,reverse=True):result=result[:st]+r+result[en:];ch.protected.append({"Input File":input_file,"Protected Type":"Model column name","Object Name":good,"Action":"Restored","Reason":f"Prevented rename from {bad}"})
    final={l:n for l,n,st,en,z in identity_blocks(result) if l};viol=[f"{n}->{final.get(l)}" for l,n,st,en,z in identity_blocks(orig) if l and final.get(l)!=n]
    if viol:raise ValueError("Model-facing name invariant failed: "+'; '.join(viol))
    return result

def format_tmdl(m):
    """Create one complete TMDL View script for each changed table.

    The complete migrated table definition retains all column settings,
    calculated columns, format strings, summarization settings, lineage tags,
    variations, annotations, and the migrated partition.
    """
    if TMDL_OUTPUT_MODE=="pbip_definition":
        return m

    definition=m.lstrip("\ufeff\r\n").replace("\r\n","\n").replace("\r","\n")
    first_content=next(
        (line.strip() for line in definition.splitlines() if line.strip()),
        "",
    )
    if not first_content.casefold().startswith("table "):
        raise ValueError(
            "TMDL View output expected an individual table definition whose "
            "first content line begins with 'table '."
        )

    indented="\n".join(
        "    "+line if line.strip() else ""
        for line in definition.split("\n")
    ).rstrip()

    return "createOrReplace\n\n"+indented+"\n"

def tmdl_data_types_by_lineage(text):
    """Return lineageTag -> dataType for physical model columns."""
    result={}
    for start,end,_model_column in col_blocks(text):
        block=text[start:end]
        lineage=re.search(r'(?m)^\s*lineageTag:\s*(\S+)\s*$',block)
        dtype=re.search(r'(?m)^\s*dataType:\s*(\S+)\s*$',block)
        if lineage and dtype:
            result[lineage.group(1)]=dtype.group(1)
    return result


def enforce_native_tmdl_type_invariant(original,migrated,input_file,ch):
    """Restore and verify native semantic dataTypes by stable lineageTag."""
    if classify_m(original)!="Native SQL":
        return migrated

    result=migrated
    original_types=tmdl_data_types_by_lineage(original)

    # Work bottom-up so text offsets remain valid.
    replacements=[]
    for start,end,model_column in col_blocks(result):
        block=result[start:end]
        lineage=re.search(r'(?m)^\s*lineageTag:\s*(\S+)\s*$',block)
        dtype=re.search(r'(?m)^\s*dataType:\s*(?P<value>\S+)\s*$',block)
        if not lineage or not dtype:
            continue
        expected=original_types.get(lineage.group(1))
        if expected and key(dtype.group('value'))!=key(expected):
            replacements.append((
                start+dtype.start('value'),
                start+dtype.end('value'),
                expected,
                model_column,
                dtype.group('value'),
            ))

    for start,end,expected,model_column,changed_type in sorted(
        replacements,reverse=True
    ):
        result=result[:start]+expected+result[end:]
        ch.protected.append({
            "Input File":input_file,
            "Protected Type":"Native TMDL dataType",
            "Object Name":model_column,
            "Count":1,
            "Action":"Restored",
            "Reason":(
                f"Prevented native semantic type change from {changed_type} "
                f"to {expected}"
            ),
        })

    final_types=tmdl_data_types_by_lineage(result)
    violations=[
        f"{lineage}: {expected} -> {final_types.get(lineage,'<missing>')}"
        for lineage,expected in original_types.items()
        if key(final_types.get(lineage,""))!=key(expected)
    ]
    if violations:
        raise ValueError(
            "NATIVE_TMDL_TYPE_INVARIANT_FAILED: "+"; ".join(violations)
        )
    return result


def process_tmdl(p,mapx,log):
    out=CFG.tmdl_out/f"{p.stem}_{CFG.direction}.tmdl"
    ch=Changes()
    original,newline_style=read_text(p)
    flushed=False
    try:
        context=infer_tables(original,mapx)|primary_tables(original)
        native=classify_m(original)=="Native SQL"
        migrated=original
        for span_start,span_end in reversed(source_spans(migrated)):
            block=migrated[span_start:span_end]
            migrated_block=(
                process_native_m(block,mapx,ch,p.name)
                if classify_m(block)=="Native SQL"
                else process_non_native(block,mapx,ch,p.name)
            )
            migrated=migrated[:span_start]+migrated_block+migrated[span_end:]
        migrated=update_tmdl_cols(migrated,mapx,context,ch,p.name,native)
        migrated=enforce_native_tmdl_type_invariant(original,migrated,p.name,ch)
        migrated=restore_names(original,migrated,ch,p.name)
        ch.changed=migrated!=original
        ch.classification=classify_m(original)
        flush(ch,log);flushed=True
        write_result(
            p,out,format_tmdl(migrated) if ch.changed else migrated,
            newline_style,ch,"TMDL",log
        )
    except Exception as exc:
        message=str(exc)
        if not flushed:
            # Preserve diagnostics collected before the failing stage, including
            # date-function conversions and SQL mapping decisions.
            flush(ch,log);flushed=True
        print_progress("TMDL",p,"Failed",message)
        log.add("10_Errors",**{
            "Input File":p.name,"Stage":"TMDL processing",
            "Error Code":type(exc).__name__,"Error Message":message,
            "Existing Output Preserved":"Yes","Processing Continued":"Yes",
        })
        log.add("2_Files Processed",**{
            "Input File":p.name,"Input Path":str(p),"File Type":"TMDL",
            "Classification":"Failed","Eligible":"Unknown",
            "Relevant Mapping Found":"Yes" if ch.relevant else "Unknown",
            "Changed":"No","Output Existed Before Run":"Yes" if out.exists() else "No",
            "Output Action":"Failed","Output File":out.name,"Output Path":str(out),
            "Write Status":"Failed","Message":message,
        })


def process_pq(p,mapx,log):
    out=p.with_name(f"{p.stem}_{CFG.direction}.txt")
    ch=Changes()
    original,newline_style=read_text(p)
    flushed=False
    try:
        migrated=(
            process_native_m(original,mapx,ch,p.name)
            if classify_m(original)=="Native SQL"
            else process_non_native(original,mapx,ch,p.name)
        )
        ch.changed=migrated!=original
        ch.classification=classify_m(original)
        flush(ch,log);flushed=True
        write_result(p,out,migrated,newline_style,ch,"Power Query TXT",log)
    except Exception as exc:
        message=str(exc)
        if not flushed:
            flush(ch,log);flushed=True
        print_progress("Power Query TXT",p,"Failed",message)
        log.add("10_Errors",**{
            "Input File":p.name,"Stage":"Power Query processing",
            "Error Code":type(exc).__name__,"Error Message":message,
            "Existing Output Preserved":"Yes","Processing Continued":"Yes",
        })
        log.add("2_Files Processed",**{
            "Input File":p.name,"Input Path":str(p),"File Type":"Power Query TXT",
            "Classification":"Failed","Eligible":"Unknown",
            "Relevant Mapping Found":"Yes" if ch.relevant else "Unknown",
            "Changed":"No","Output Existed Before Run":"Yes" if out.exists() else "No",
            "Output Action":"Failed","Output File":out.name,"Output Path":str(out),
            "Write Status":"Failed","Message":message,
        })


# =============================================================================
# EXCEL LOG AND MAIN
# =============================================================================
def style_log(path:Path):
    wb=load_workbook(path)
    header_fill=PatternFill("solid",fgColor="1F4E78");header_font=Font(color="FFFFFF",bold=True)
    green=PatternFill("solid",fgColor="C6EFCE");yellow=PatternFill("solid",fgColor="FFF2CC")
    red=PatternFill("solid",fgColor="FFC7CE");blue=PatternFill("solid",fgColor="DDEBF7")
    try:
        for ws in wb.worksheets:
            ws.freeze_panes="A2"
            for c in ws[1]: c.fill=header_fill;c.font=header_font;c.alignment=Alignment(horizontal="center",vertical="center")
            for row in ws.iter_rows(min_row=2):
                for c in row:
                    c.alignment=Alignment(vertical="top",wrap_text=True)
                    v=str(c.value or "").casefold()
                    if v in {"created","overwritten","unchanged","success","completed"}:c.fill=green
                    elif v in {"skipped","stale output retained","completed with warnings"}:c.fill=yellow
                    elif v in {"failed","completed with errors"}:c.fill=red
                    elif v in {"would create","would overwrite","dry run"}:c.fill=blue
            for i in range(1,ws.max_column+1):
                vals=[str(ws.cell(r,i).value or "") for r in range(1,min(ws.max_row,250)+1)]
                ws.column_dimensions[get_column_letter(i)].width=min(55,max(12,max((len(v) for v in vals),default=0)+2))
            if ws.max_row>=2 and ws.max_column>=1:
                ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                tab=Table(displayName=("Log"+re.sub(r'\W','',ws.title))[:250],ref=ref)
                tab.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showFirstColumn=False,showLastColumn=False,showColumnStripes=False)
                ws.add_table(tab)
        wb.save(path)
    finally: wb.close()


def write_log(log:Log,tcount,qcount):
    files=log.rows["2_Files Processed"]
    actions=[r.get("Output Action","") for r in files]
    warning_count=len(log.rows["9_Warnings"]);error_count=len(log.rows["10_Errors"])
    overall="Completed with errors" if error_count else "Completed with warnings" if warning_count else "Completed"
    vals=[
        ("Run ID",log.rid),("Run Timestamp",log.timestamp),("Direction",CFG.direction.title()),("Dry Run","Yes" if CFG.dry_run else "No"),
        ("Mapping Workbook",str(CFG.workbook)),("Mapping Sheet",CFG.sheet),("TMDL Input Folder",str(CFG.tmdl_in)),
        ("TMDL Output Folder",str(CFG.tmdl_out)),("TMDL Output Mode",TMDL_OUTPUT_MODE),
        ("Disabled Query Folder",str(CFG.pq_folder or "Not supplied")),("Disabled Query Processing","Enabled" if CFG.pq_folder else "Skipped"),
        ("TMDL Files Scanned",tcount),("Disabled Query Files Scanned",qcount),("Outputs Created",actions.count("Created")),
        ("Outputs Overwritten",actions.count("Overwritten")),("Outputs Unchanged",actions.count("Unchanged")),
        ("Would Create",actions.count("Would create")),("Would Overwrite",actions.count("Would overwrite")),
        ("Files Skipped",actions.count("Skipped")),("Stale Outputs Retained",actions.count("Stale output retained")),
        ("Warnings",warning_count),("Errors",error_count),("Legacy Retention Rules Configured",len(LEGACY_TYPE_RETENTION_RULES)),
        ("Per-Column Overrides Configured",len(COLUMN_TYPE_OVERRIDES)),("Original PBIP Files Modified","No"),
        ("Original Query TXT Modified","No"),("Model-Facing Names Modified","No"),("DAX Expressions Modified","No"),
        ("Overall Status",overall)]
    log.rows["1_Run Summary"]=[{"Item":a,"Value":b} for a,b in vals]
    final=CFG.log_out/LOG_FILENAME
    temp=None
    try:
        with tempfile.NamedTemporaryFile(mode="wb",dir=CFG.log_out,prefix="migration_log_",suffix=".tmp.xlsx",delete=False) as f: temp=Path(f.name)
        with pd.ExcelWriter(temp,engine="openpyxl") as w:
            for sheet,cols in SHEETS.items(): pd.DataFrame(log.rows[sheet],columns=cols).to_excel(w,sheet_name=sheet,index=False)
        style_log(temp)
        try: os.replace(temp,final)
        except PermissionError as e:
            raise PermissionError(f"Could not update {final}. Close the workbook in Excel or File Explorer Preview and rerun.") from e
        temp=None
        return final
    finally:
        if temp and temp.exists():
            try: temp.unlink()
            except OSError: pass


def print_run_snapshot(log:Log,tcount:int,qcount:int,log_path:Path):
    actions=defaultdict(int)
    for row in log.rows["2_Files Processed"]:actions[row.get("Output Action","Unknown")]+=1
    print("\n"+"="*68)
    print("POWER BI SOURCE MIGRATION RUN SNAPSHOT")
    print("="*68)
    print(f"Run ID:                       {log.rid}")
    print(f"Direction:                    {CFG.direction}")
    print(f"Dry run:                      {CFG.dry_run}")
    print(f"TMDL files scanned:           {tcount}")
    print(f"Disabled query files scanned: {qcount}")
    for name in ("Created","Overwritten","Unchanged","Would create","Would overwrite","Skipped","Stale output retained","Failed"):
        if actions[name]:print(f"{name+':':30}{actions[name]}")
    print(f"Warnings:                     {len(log.rows['9_Warnings'])}")
    print(f"Errors:                       {len(log.rows['10_Errors'])}")
    print("Original PBIP files modified: No")
    print("Model-facing names modified:  No")
    print(f"Migration log:                {log_path}")
    generated_count = sum(actions[name] for name in ("Created", "Overwritten", "Unchanged"))
    if generated_count == 0:
        print("-")
        if CFG.dry_run:
            print("No TMDL files were written because DRY_RUN=True.")
            print("Set DRY_RUN=False after reviewing Would create/Would overwrite entries.")
        elif actions["Failed"]:
            print("No TMDL files were written because file processing failed.")
            print("Review 10_Errors and the Failed rows in 2_Files Processed.")
        else:
            print("No TMDL files were written because no mapped source changes were detected.")
            print("Review Relevant Mapping Found and Message in 2_Files Processed.")
    print("="*68)


def parse_cli_args(argv=None):
    parser=argparse.ArgumentParser(
        description="Migrate Power BI TMDL and disabled Power Query sources using a legacy-to-target mapping workbook."
    )
    parser.add_argument("--mapping-workbook",default=MAPPING_WORKBOOK)
    parser.add_argument("--mapping-sheet",default=MAPPING_SHEET)
    parser.add_argument("--tmdl-input",default=TMDL_TABLES_INPUT_FOLDER)
    parser.add_argument("--tmdl-output",default=TMDL_OUTPUT_FOLDER)
    parser.add_argument("--log-output",default=LOG_OUTPUT_FOLDER)
    parser.add_argument("--disabled-query-folder",default=DISABLED_POWER_QUERY_FOLDER)
    parser.add_argument("--direction",choices=("target","legacy"),default=DIRECTION)
    mode=parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run",action="store_true",help="Analyze and log without writing migrated outputs")
    mode.add_argument("--write",action="store_true",help="Write migrated outputs")
    parser.add_argument("--no-console-progress",action="store_true")
    return parser.parse_args(argv)


def apply_cli_args(args):
    global DISABLED_POWER_QUERY_FOLDER,MAPPING_WORKBOOK,MAPPING_SHEET
    global TMDL_TABLES_INPUT_FOLDER,TMDL_OUTPUT_FOLDER,LOG_OUTPUT_FOLDER
    global DIRECTION,DRY_RUN,LIVE_CONSOLE_PROGRESS
    DISABLED_POWER_QUERY_FOLDER=args.disabled_query_folder or ""
    MAPPING_WORKBOOK=args.mapping_workbook
    MAPPING_SHEET=args.mapping_sheet
    TMDL_TABLES_INPUT_FOLDER=args.tmdl_input
    TMDL_OUTPUT_FOLDER=args.tmdl_output
    LOG_OUTPUT_FOLDER=args.log_output
    DIRECTION=args.direction
    if args.dry_run:
        DRY_RUN=True
    elif args.write:
        DRY_RUN=False
    LIVE_CONSOLE_PROGRESS=not args.no_console_progress


def main(argv=None):
    global CFG,QUERY_CONTEXT
    apply_cli_args(parse_cli_args(argv))
    CFG=build_config();validate(CFG);maps=load_mapping(CFG)
    rid=datetime.now().strftime("%Y%m%d_%H%M%S")+"_"+CFG.direction;log=Log(rid)
    pq=discover_pq(CFG);QUERY_CONTEXT=build_context(pq,maps)
    tmdl=sorted(CFG.tmdl_in.glob("*.tmdl"),key=lambda p:p.name.casefold())
    print("\nPower BI source migration started")
    print(f"Direction: {CFG.direction} | Dry run: {CFG.dry_run}")
    print(f"TMDL input: {CFG.tmdl_in}")
    print(f"TMDL output: {CFG.tmdl_out}")
    print(f"Disabled queries: {CFG.pq_folder or 'Not supplied'}\n")
    for p in pq:process_pq(p,maps,log)
    for p in tmdl:process_tmdl(p,maps,log)
    lp=write_log(log,len(tmdl),len(pq));print_run_snapshot(log,len(tmdl),len(pq),lp)
    return 0 if not log.rows["10_Errors"] else 2
if __name__=="__main__":raise SystemExit(main())
